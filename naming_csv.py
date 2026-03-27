#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
naming_csv.py — KRUK / PLAZAS ESPECIALES

Para cada expediente en IN/ genera las filas del CSV de mapeo:
  referencia_demanda | nombre_fichero_original | nombre_correcto | numero_documento | status | motivo

Uso:
  python3 naming_csv.py [ROOT_DIR] [EXCEL_REGLAS]

Defaults:
  ROOT_DIR     = /Users/sergio.garcia/Desktop/DEMANDAS
  EXCEL_REGLAS = ROOT_DIR/IN/reglas renombrado.xlsx
  IN_DIR       = ROOT_DIR/IN
  INDICES_DIR  = ROOT_DIR/IN/iNDICES
"""

import os
import re
import csv
import sys
import unicodedata
from typing import Dict, List, Optional, Tuple

# ──────────────────────────────────────────
# VENV AUTO-ACTIVACIÓN
# ──────────────────────────────────────────

def _ensure_venv() -> None:
    import platform as _platform
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    if _platform.system() == "Windows":
        _venv_python = os.path.join(_script_dir, "venv", "Scripts", "python.exe")
    else:
        _venv_python = os.path.join(_script_dir, "venv", "bin", "python3")
    if os.path.isfile(_venv_python):
        if os.path.realpath(sys.executable) != os.path.realpath(_venv_python):
            print(f"[venv] Re-ejecutando con: {_venv_python}")
            if _platform.system() == "Windows":
                import subprocess as _sp
                sys.exit(_sp.run([_venv_python] + sys.argv).returncode)
            else:
                os.execv(_venv_python, [_venv_python] + sys.argv)

_ensure_venv()

if sys.version_info < (3, 6):
    sys.exit(f"[ERROR] Se requiere Python 3.6 o superior. Versión actual: {sys.version}")

# ──────────────────────────────────────────
# AUTO-INSTALACIÓN DE DEPENDENCIAS
# ──────────────────────────────────────────

def _ensure_deps() -> None:
    _deps = [
        ("fitz",     "PyMuPDF"),
        ("pypdf",    "pypdf"),
        ("openpyxl", "openpyxl"),
        ("pymysql",  "pymysql"),
    ]
    import importlib.util, subprocess
    for module, package in _deps:
        if importlib.util.find_spec(module) is None:
            print(f"[deps] Instalando {package} …")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])

_ensure_deps()

# ──────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────

# ── Base de datos Kmaleon ──────────────────────────────────────
DB_HOST  = "172.26.3.45"
DB_PORT  = 3306
DB_USER  = "sys.etl"
DB_PASS  = "oofu6Ahs9ix4uiloot6d"
DB_NAME  = "kmaleon"

# Orden de asignación posicional de CTFDOs (email=1º, sms=2º, burofax=3º)
_CHANNEL_ORDER = ["EMAIL", "SMS", "BUROFAX"]

# Documentos comunes ubicados en ROOT (DEMANDAS/) compartidos por todos los expedientes.
# (nombre_fichero, idx_kw para buscar el número en el índice)
_COMMON_DOCS: List[Tuple[str, str]] = [
    ("CERTIFICADO DE TITULARIDAD CUENTA.pdf",    "Certificado de titularidad"),
    ("DOC.1 PODER INVESTCAPITAL.pdf",            "Poder general para pleitos"),
]

# Doc común que se incluye en demandas.xlsx SOLO cuando el índice del expediente
# contiene "Escritura de fusión". (fn_fichero, idx_kw)
_NOTARIAL_COMMON_DOC: Tuple[str, str] = (
    "TESTIMONIO FUSION BANKIA-CAIXA.pdf",
    "Escritura de fusión",
)

# Expansión de tokens cortos usados en col A del Excel
_TOKEN_EXPAND = {"cont": "contenido"}

# Reglas de nomenclatura hardcodeadas (col A → col B → col C del Excel original)
# Ordenadas de más específica a menos (mayor longitud de fn_kw primero).
_RULES: List[Tuple[str, str, str]] = sorted([
    # fn_kw                      idx_kw                                      template
    ("CONTRATO",               "Contrato suscrito por el titular",            "DOC. X CONTRATO"),
    ("MOVIMIENTOS",            'Que contenga "movimientos"',                  "DOC. X MOVIMIENTOS"),
    ("PODER",                  "Poder general para pleitos",                  "DOC. X PODER"),
    ("TESTIMONIO",             "Testimonio notarial",                         "DOC. X TESTIMONIO"),
    ("sms cont",               "Certificado de contenido.",                   "DOC. X CERTIFICADO CONTENIDO SMS"),
    ("sms",                    "Justificante de intento de notificación",     "DOC. X CERTIFICADO FEHACIENTE SMS"),
    ("titularidad bancaria",   "Certificado de titularidad de la cuenta",     "DOC. X CERTIFICADO TITULARIDAD"),
    ("ctfdo deuda",            "Certificado saldo",                           "DOC. X CERTIFICADO SALDO"),
    ("burofax cont",           "Certificados de contenido",                   "DOC. X CERTIFICADO CONTENIDO BUROFAX"),
    ("email cont",             "Certificados de contenido",                   "DOC. X CERTIFICADO CONTENIDO EMAIL"),
    ("email",                  "Justificante de intento de notificación",     "DOC. X CERTIFICADO FEHACIENTE EMAIL"),
    ("burofax",                "Justificante de intento de notificación",     "DOC. X CERTIFICADO FEHACIENTE BUROFAX"),
    ("Tasa",                   "abono tasa",                                  "DOC. X TASA"),
    ("Tasa",                   "Justificante tasa judicial",                  "DOC. X TASA"),
    ("decl responsable",       "Declaración de responsabilidad",              "DOC. X DECLARACIÓN RESPONSABLE"),
    ("ctfdo deuda IC",         "Certificado de deuda",                        "DOC. X CERTIFICADO IC"),
    ("hello letter",           "Notificación de la cesión",                   "DOC. X NOTIFICACIÓN DE LA CESION"),
    ("welcome letter",         "Notificación de la cesión",                   "DOC. X NOTIFICACIÓN DE LA CESION"),
    ("movimientos 1",          "Movimientos de origen de la deuda",           "DOC. X MOVIMIENTOS"),
    ("movimientos 1",          "Liquidación tras el pase a mora",             "DOC. X MOVIMIENTOS"),  # fallback si solo hay este tipo en el índice
    ("movimientos 2",          "Liquidación tras el pase a mora",             "DOC. X MOVIMIENTOS 2"),
    ("movimientos 2",          "Movimientos de origen de la deuda",           "DOC. X MOVIMIENTOS 2"),  # fallback si solo hay este tipo en el índice
    # Aliases para ficheros con nomenclatura alternativa (ej. _firmado_TIPO)
    ("certificado ic",         "Certificado de deuda",                        "DOC. X CERTIFICADO IC"),
    ("declaracion responsable","Declaración de responsabilidad",              "DOC. X DECLARACIÓN RESPONSABLE"),
    ("declaracion responsable","Declaracion responsable",                     "DOC. X DECLARACIÓN RESPONSABLE"),
    ("certificado saldo",      "Certificado saldo",                           "DOC. X CERTIFICADO SALDO"),
    ("certificado correspondencia", "Certificado de correspondencia numérica.", "DOC. X CERTIFICADO CORRESPONDENCIA"),
    ("firmado acuerdo",            "Convenio de amortización y reconocimiento de deuda", "DOC. X ACUERDO"),
    ("firmado acuerdo",            "Documento acreditativo actividad negociadora previa", "DOC. X ACUERDO"),
], key=lambda r: len(r[0]), reverse=True)

# ──────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────

def norm(s: str) -> str:
    """Colapsa espacios y elimina NBSP."""
    s = (s or "").replace("\u00a0", " ")
    return re.sub(r"\s+", " ", s).strip()

# Caracteres Unicode que el OCR confunde con letras ASCII comunes.
# Se aplican ANTES de la normalización NFKD para que no pasen al índice.
_OCR_CHAR_MAP = str.maketrans({
    # Variantes de 'l' (ele minúscula)
    "\u01C0": "l",   # ǀ Latin letter dental click
    "\u0269": "l",   # ɩ Latin small letter iota
    "\u2113": "l",   # ℓ Script small l
    "\u006C": "l",   # l  (propio, por si acaso)
    "\uFF4C": "l",   # ｌ Fullwidth small l
    # Variantes de 'I' (i mayúscula) que pueden confundirse con l
    "\u0406": "i",   # І Cyrillic I
    "\u04C0": "i",   # Ӏ Modifier letter capital i (Checheno)
    # Variantes de 'o'
    "\u00BA": "o",   # º Ordinal masculino (confundible en fuentes OCR)
})

def al(s: str) -> str:
    """ASCII-lower: normaliza unicode, elimina tildes, minúsculas, colapsa espacios.
    Aplica correcciones de caracteres OCR problemáticos antes de la normalización."""
    s = (s or "").translate(_OCR_CHAR_MAP)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.lower()).strip()

# ──────────────────────────────────────────
# MAPAS DE CLASE
# ──────────────────────────────────────────

# Mapeo fn_kw → valor de la columna Clase en naming_output.csv
_CLASE_MAP: Dict[str, str] = {
    al("Tasa"):                                         "TASA",
    al("ctfdo deuda IC"):                               "CERTIFICADO",
    al("certificado ic"):                               "CERTIFICADO",
    al("CONTRATO"):                                     "CONTRATO",
    al("burofax cont"):                                 "BUROFAX",
    al("sms cont"):                                     "SMS",
    al("burofax"):                                      "JUST BUROFAX",
    al("sms"):                                          "JUST SMS",
    al("decl responsable"):                             "CERTIFICADO_1",
    al("declaracion responsable"):                      "CERTIFICADO_1",
    al("hello letter"):                                 "HELLO LETTER",
    al("welcome letter"):                               "HELLO LETTER",
    al("MOVIMIENTOS"):                                  "MOVIMIENTOS",
    al("movimientos 1"):                                "MOVIMIENTOS",
    al("movimientos 2"):                                "MOVIMIENTOS 2",
    al("TESTIMONIO"):                                   "TESTIMONIO",
    al("titularidad bancaria"):                         "CERTIF",
    al("Certificado de titularidad de la cuenta"):      "CERTIF",   # doc común
    al("Certificado de titularidad"):                   "CERTIF",   # doc común (idx_kw corto)
    al("email cont"):                                   "EMAIL",
    al("email"):                                        "JUST EMAIL",
    al("ctfdo deuda"):                                  "CERTIF_1",
    al("certificado saldo"):                            "CERTIF_1",
    al("PODER"):                                        "PODER",
    al("Poder general para pleitos"):                   "PODER",    # doc común
    al("certificado correspondencia"):                  "CERT_COR_NUM",
    al("INDICE"):                                       "INDICE",
    al("firmado acuerdo"):                              "ACUERDO",
    al("Escritura de fusión"):                          "NOTARIAL",  # doc común
}

# Mapeo fn_kw → Clase para procedimiento monitorio en plaza especial
_CLASE_MAP_MON: Dict[str, str] = {
    al("Tasa"):                                         "TASA MON",
    al("ctfdo deuda IC"):                               "CERTIFICADO MON",
    al("certificado ic"):                               "CERTIFICADO MON",
    al("CONTRATO"):                                     "CONTRATO MON",
    al("burofax cont"):                                 "BUROFAX MON",
    al("sms cont"):                                     "SMS MON",
    al("burofax"):                                      "JUST BUROFAX MON",
    al("sms"):                                          "JUST SMS MON",
    al("decl responsable"):                             "CERTIF MON",
    al("declaracion responsable"):                      "CERTIF MON",
    al("MOVIMIENTOS"):                                  "EXTRACTO MON",
    al("movimientos 1"):                                "EXTRACTO MON",
    al("movimientos 2"):                                "MOVIMIENTOS 2 MON",
    al("TESTIMONIO"):                                   "TESTIMONIO MON",
    al("titularidad bancaria"):                         "CERTIF DEUDA MON",
    al("Certificado de titularidad de la cuenta"):      "CERTIF DEUDA MON",  # doc común
    al("Certificado de titularidad"):                   "CERTIF DEUDA MON",  # doc común (idx_kw corto)
    al("email cont"):                                   "EMAIL MON",
    al("email"):                                        "JUST EMAIL MON",
    al("ctfdo deuda"):                                  "CERTIF SALDO MON",
    al("certificado saldo"):                            "CERTIF SALDO MON",
    al("PODER"):                                        "PODER MON",
    al("Poder general para pleitos"):                   "PODER MON",  # doc común
    al("INDICE"):                                       "INDICE MON",
    al("firmado acuerdo"):                              "ACUERDO MON",
    al("Escritura de fusión"):                          "NOTARIAL MON",  # doc común
}

# Ciudades que, en procedimiento monitorio, usan _CLASE_MAP_MON
_MONITORIO_EXCLUSION_CITIES_NORM: set = {al(c) for c in [
    "Albacete", "Alcaraz", "Alcázar de San Juan", "Almadén", "Almagro",
    "Almansa", "Almazán", "Almendralejo", "Aranda de Duero",
    "Arenas de San Pedro", "Arévalo", "Astorga", "Ávila",
    "Badajoz", "La Bañeza", "Béjar", "Benavente", "Briviesca",
    "Burgo de Osma-Ciudad de Osma", "Burgos", "Cáceres",
    "Caravaca de la Cruz", "Carrión de los Condes", "Cartagena",
    "Casas-Ibáñez", "Castuera", "Cervera de Pisuerga", "Ceuta",
    "Cieza", "Cistierna", "Ciudad Real", "Ciudad Rodrigo",
    "Ciutadella de Menorca", "Coria", "Cuéllar", "Cuenca",
    "Daimiel", "Don Benito", "Eivissa", "Fregenal de la Sierra",
    "Guadalajara", "Hellín", "Herrera del Duque", "Illescas",
    "Inca", "Jerez de los Caballeros", "Jumilla", "León", "Lerma",
    "Llerena", "Logrosán", "Lorca", "Manacor", "Manzanares",
    "Mahón", "Medina de Rioseco", "Medina del Campo", "Melilla",
    "Mérida", "Miranda de Ebro", "Molina de Aragón", "Molina de Segura",
    "Montijo", "Motilla del Palancar", "Mula", "Murcia",
    "Navalmoral de la Mata", "Ocaña", "Olivenza", "Orgaz",
    "Palencia", "Palma de Mallorca", "Peñaranda de Bracamonte",
    "Piedrahíta", "Plasencia", "Ponferrada", "Puebla de Sanabria",
    "Puertollano", "Quintanar de la Orden", "La Roda", "Sahagún",
    "Salamanca", "Salas de los Infantes", "San Clemente", "San Javier",
    "Santa María la Real de Nieva", "Segovia", "Sepúlveda", "Sigüenza",
    "Soria", "Talavera de la Reina", "Tarancón", "Toledo", "Tomelloso",
    "Toro", "Torrijos", "Totana", "Trujillo", "Valdepeñas",
    "Valencia de Alcántara", "Valladolid", "Villablino",
    "Villafranca de los Barros", "Villalpando",
    "Villanueva de la Serena", "Villanueva de los Infantes",
    "Villarcayo de Merindad de Castilla la Vieja", "Villarrobledo",
    "Vitigudino", "Yecla", "Zafra", "Zamora",
]}


def _get_clase(fn_kw: str, tipo_proc: str, city: str) -> str:
    """
    Devuelve el valor de Clase para naming_output.csv.
    - Monitorio + plaza especial  → _CLASE_MAP_MON
    - Resto (incluido monitorio en plaza no especial) → _CLASE_MAP
    """
    key = al(fn_kw)
    if "monitorio" in al(tipo_proc) and al(city) in _MONITORIO_EXCLUSION_CITIES_NORM:
        return _CLASE_MAP_MON.get(key, "")
    return _CLASE_MAP.get(key, "")

# ──────────────────────────────────────────
# REGLAS CSV
# ──────────────────────────────────────────


def load_datatape(path: str) -> Dict[str, Dict]:
    """
    Lee datatape.xlsx y devuelve dict: exp → {idx_num, referencia2, tipo_proc, city}.
    Lee las columnas por NOMBRE para que sea robusto ante reordenaciones:
      'Whole Case Number'       → referencia  (query kmaleon)
      'Original Contract Number'→ exp / referencia2 (clave del expediente)
      'City Court'              → city
      'Tipo procedimiento'      → tipo_proc
      'indice'                  → número de índice
    """
    import openpyxl
    out: Dict[str, Dict] = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # ── Leer cabecera y construir mapa nombre→índice ──────────────
    headers = {}
    header_row = next(ws.iter_rows(values_only=True))
    for i, cell in enumerate(header_row):
        if cell is not None:
            headers[str(cell).strip()] = i

    _COL_EXP   = "Original Contract Number"
    _COL_REF2  = "Whole Case Number"
    _COL_CITY  = "City Court"
    _COL_TIPO  = "Tipo procedimiento"
    _COL_IDX   = "Indice"

    missing = [c for c in [_COL_EXP, _COL_REF2, _COL_CITY, _COL_TIPO, _COL_IDX]
               if c not in headers]
    if missing:
        print(f"[datatape] WARN columnas no encontradas: {missing}")

    def _get(row, col_name):
        idx = headers.get(col_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def _cell_str(val) -> str:
        """Convierte celda a string limpio. Floats enteros → int (evita '4511220000929.0')."""
        if val is None:
            return ""
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val).strip()

    # ── Leer filas de datos ───────────────────────────────────────
    for row in ws.iter_rows(min_row=2, values_only=True):
        exp     = _cell_str(_get(row, _COL_EXP))
        ref2    = _cell_str(_get(row, _COL_REF2))
        city    = str(_get(row, _COL_CITY) or "").strip()
        tipo    = str(_get(row, _COL_TIPO) or "").strip()
        idx_raw = _get(row, _COL_IDX)
        if not exp or exp == "None":
            continue
        try:
            idx_num = int(float(str(idx_raw)))
        except (TypeError, ValueError):
            idx_num = None
        out[exp] = {"idx_num": idx_num, "referencia2": ref2, "tipo_proc": tipo, "city": city}

    wb.close()
    print(f"[datatape] {len(out)} expedientes cargados desde: {os.path.basename(path)}")
    return out


def load_kmaleon_codes(datatape: Dict[str, Dict]) -> Dict[str, str]:
    """
    Consulta Kmaleon y devuelve dict: exp (col C) → Asunto_Codigo.
    Por cada expediente usa:
      referencia      = exp (col C, Original Contract Number)
      referencia2     = col B (Whole Case Number)
      tipoprocedimiento LIKE %<col K>%
    Ejecuta una query por expediente reutilizando la misma conexión.
    """
    import pymysql
    if not datatape:
        return {}

    sql = """
        SELECT
            CASE WHEN asu.codserie = 1 THEN ''
                 ELSE CONCAT(asu.serie, '-')
            END
            || asu.codigo
            || CASE WHEN asu.subcodigo = 0 THEN ''
                    ELSE CONCAT('.', LPAD(asu.subcodigo, 3, '0'))
               END AS Asunto_Codigo
        FROM asuntos asu
        INNER JOIN asunorganos ong
               ON asu.interno = ong.codigoasunto AND ong.linea = 1
        WHERE asu.referencia       = %s
          AND asu.referencia2      = %s
          AND ong.tipoprocedimiento LIKE %s
    """
    result: Dict[str, str] = {}
    try:
        conn = pymysql.connect(
            host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
            database=DB_NAME, charset="utf8mb4",
        )
        dup_count = 0
        with conn.cursor() as cur:
            for exp, info in datatape.items():
                referencia  = info.get("referencia2", "")  # col B = ICARR/1/XXXX
                referencia2 = exp                           # col C = número expediente
                tipo        = info.get("tipo_proc", "")
                if not referencia or not tipo:
                    continue
                cur.execute(sql, (referencia, referencia2, f"%{tipo}%"))
                rows = cur.fetchall()
                # Si no encuentra resultado y el tipo es hipotecario,
                # reintenta buscando por 'hipotecaria'
                if not rows and "hipotecario" in tipo.lower():
                    cur.execute(sql, (referencia, referencia2, "%hipotecaria%"))
                    rows = cur.fetchall()
                if len(rows) > 1:
                    dup_count += 1
                    codigos = [str(r[0]) for r in rows if r[0] is not None]
                    print(f"[kmaleon] WARN expediente '{exp}' devuelve {len(rows)} resultados: "
                          f"{', '.join(codigos)} → se usa el primero ({codigos[0] if codigos else '-'})")
                if rows and rows[0][0] is not None:
                    result[exp] = str(rows[0][0])
        conn.close()
        print(f"[kmaleon] {len(result)}/{len(datatape)} códigos obtenidos")
        if dup_count == 0:
            print("[kmaleon] No encontrados expedientes duplicados en kmaleon")
    except Exception as e:
        print(f"[kmaleon] ERROR conectando a la BD: {e}")
    return result


def apply_x(template: str, x: int) -> str:
    """Sustituye la primera aparición de 'DOC. X ...' por 'DOC. N ...'"""
    return re.sub(r"(\bDOC\.\s*)X\b", rf"\g<1>{x}", template,
                  count=1, flags=re.IGNORECASE)

# ──────────────────────────────────────────
# EXTRACCIÓN DE TEXTO PDF
# ──────────────────────────────────────────

def pdf_text(path: str) -> str:
    """Extrae texto del PDF (fitz primero, pypdf como fallback)."""
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(path)
        t = "\n".join(p.get_text() for p in doc)
        doc.close()
        return t
    except Exception:
        pass
    try:
        from pypdf import PdfReader
        return "\n".join(p.extract_text() or "" for p in PdfReader(path).pages)
    except Exception:
        return ""


def extract_last_page_as_index(exp_dir: str, exp: str,
                               exp_idx: int = 0, exp_total: int = 0) -> Optional[str]:
    """
    Busca el PDF de demanda en exp_dir, extrae su última página
    y la guarda como INDICE_{exp}.pdf en la misma carpeta.
    Estrategia de búsqueda (igual que collect_demandas):
      1. Primer PDF cuyo nombre contiene 'demanda'
      2. Fallback: {exp}_firmado.pdf
    Devuelve la ruta al fichero creado, o None si no hay demanda.
    """
    all_pdfs = sorted(f for f in os.listdir(exp_dir) if f.lower().endswith(".pdf"))

    demanda_pdf: Optional[str] = None

    # 1. Buscar por nombre con 'demanda'
    for fn in all_pdfs:
        if "demanda" in al(fn):
            demanda_pdf = os.path.join(exp_dir, fn)
            break

    # 2. Fallback: {exp}_firmado.pdf (misma lógica que collect_demandas)
    if demanda_pdf is None:
        firmado_name = f"{exp}_firmado.pdf"
        for fn in all_pdfs:
            if fn.lower() == firmado_name.lower():
                demanda_pdf = os.path.join(exp_dir, fn)
                print(f"[indice] '{exp}': sin DEMANDA en nombre, usando fallback '{fn}'")
                break

    _cnt    = f"[{exp_idx}/{exp_total}]  " if exp_total else ""
    _rest   = exp_total - exp_idx
    _sufijo = f"  —  {_rest} por procesar" if _rest > 0 else "  —  último"

    if demanda_pdf is None:
        print(f"[indice] {_cnt}WARN '{exp}': no se encontró PDF de DEMANDA para extraer índice{_sufijo}")
        return None

    out_path = os.path.join(exp_dir, f"INDICE_{exp}.pdf")
    src_fn   = os.path.basename(demanda_pdf)

    try:
        import fitz
        doc = fitz.open(demanda_pdf)
        last = len(doc) - 1
        new_doc = fitz.open()
        new_doc.insert_pdf(doc, from_page=last, to_page=last)
        new_doc.save(out_path)
        new_doc.close()
        doc.close()
        print(f"[indice] {_cnt}'{exp}': índice extraído de '{src_fn}' (última pág. {last + 1}/{last + 1}){_sufijo}")
        return out_path
    except Exception as e_fitz:
        print(f"[indice] {_cnt}WARN '{exp}': fitz falló ({e_fitz}), reintentando con pypdf…")

    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(demanda_pdf)
        writer = PdfWriter()
        writer.add_page(reader.pages[-1])
        with open(out_path, "wb") as f:
            writer.write(f)
        print(f"[indice] {_cnt}'{exp}': índice extraído de '{src_fn}' (pypdf, última pág. {len(reader.pages)}/{len(reader.pages)}){_sufijo}")
        return out_path
    except Exception as e_pypdf:
        print(f"[indice] {_cnt}ERROR '{exp}': no se pudo extraer última página de '{src_fn}': "
              f"fitz={e_fitz} | pypdf={e_pypdf}{_sufijo}")
        return None

# ──────────────────────────────────────────
# PARSING DEL ÍNDICE
# ──────────────────────────────────────────

class IdxItem:
    def __init__(self, num: int, desc: str, dn: str):
        self.num  = num
        self.desc = desc   # descripción original
        self.dn   = dn     # al(desc) para comparaciones

_LINE_RE = re.compile(
    r"\bdocumento(?:s)?\b\s*(?:n\.?[ºo°]\s*)?(?P<nums>[^:\-–]{1,40}?)\s*(?::\s*|[\-–]\s+)\s*(?P<desc>.+?)\s*$",
    re.IGNORECASE,
)

def _parse_nums(s: str) -> List[int]:
    """'4, 5 y 6' → [4, 5, 6].
    Corrige confusiones OCR: 'l' / 'I' → 1, 'O' → 0."""
    s = re.sub(r"[ºo°,\.;]", " ", s)
    s = re.sub(r"\by\b", " ", s, flags=re.IGNORECASE)
    _OCR_DIGIT = {"l": 1, "i": 1, "I": 1, "O": 0}
    result = []
    for t in s.split():
        if t.isdigit():
            result.append(int(t))
        elif t in _OCR_DIGIT:
            result.append(_OCR_DIGIT[t])
    return result

def parse_index(text: str) -> List[IdxItem]:
    """Extrae items del índice y los devuelve ordenados por número."""
    items = []
    for raw in text.splitlines():
        ln = norm(raw)
        m = _LINE_RE.search(ln)
        if not m:
            continue
        nums = _parse_nums(m.group("nums"))
        desc = norm(m.group("desc")).rstrip(".")
        if not nums or not desc:
            continue
        dn = al(desc)
        for n in nums:
            items.append(IdxItem(n, desc, dn))
    return sorted(items, key=lambda x: x.num)

def find_index_pdf(idx_num: Optional[int], indices_dir: str) -> Optional[str]:
    """
    Busca en indices_dir el PDF cuyo nombre empieza por 'INDICE <idx_num>'.
    Ejemplo: idx_num=9 → 'INDICE 9 CARREFOUR - ...pdf'
    """
    if idx_num is None or not os.path.isdir(indices_dir):
        return None
    prefix = f"indice {idx_num}"
    for fn in sorted(os.listdir(indices_dir)):
        if not fn.lower().endswith(".pdf"):
            continue
        if al(fn).startswith(prefix):
            return os.path.join(indices_dir, fn)
    return None

# ──────────────────────────────────────────
# CLASIFICACIÓN POR NOMBRE DE FICHERO
# ──────────────────────────────────────────

def _channel(fn: str) -> str:
    """Detecta canal de comunicación en el nombre del fichero."""
    t = al(fn)
    if "burofax" in t:             return "BUROFAX"
    if re.search(r"\bsms\b", t):   return "SMS"
    if "email" in t:               return "EMAIL"
    return "UNKNOWN"

def classify(fn: str) -> Tuple[str, str]:
    """
    Clasifica un fichero por su nombre.
    Devuelve (doc_type, motivo).
    doc_type == 'SKIP' → no generar fila en el output.
    doc_type == 'UNKNOWN' → fila con status FAIL.
    """
    base = os.path.basename(fn)
    t = al(base)

    # ── Ficheros a ignorar ────────────────────────────────────────
    if t.startswith("indice"):
        return "SKIP", "indice"
    if "plantilla" in t and ("resumen" in t or t.strip() == "plantilla"):
        return "SKIP", "plantilla_resumen"
    if base.lower().endswith("_ocr.pdf"):
        # PODER INVESTCAPITAL_ocr.pdf: duplicado OCR, se usa el original
        return "SKIP", "ocr_dup"
    if "demanda" in t:
        return "SKIP", "demanda"

    # ── PODER ────────────────────────────────────────────────────
    # Cubre "PODER INVESTCAPITAL.PDF" y cualquier variante con 'poder'
    if "poder" in t:
        return "PODER", "fn:poder"

    # ── CTFDOs ──────────────────────────────────────────────────
    if "ctfdo" in t and "contenido" in t:
        ch = _channel(base)
        return f"CTFDO_CONTENIDO_{ch}", f"fn:ctfdo_contenido_{ch}"

    if "ctfdo" in t:
        # fehaciente / remision / recepcion — lo que quede de CTFDO
        ch = _channel(base)
        return f"CTFDO_FEHACIENTE_{ch}", f"fn:ctfdo_fehaciente_{ch}"

    # ── Documentos por nombre explícito ──────────────────────────
    if "contrato"                    in t: return "CONTRATO",                 "fn:contrato"
    if "testimonio"                  in t: return "TESTIMONIO",               "fn:testimonio"
    if "declaracion responsable"     in t: return "DECLARACION_RESPONSABLE",  "fn:decl_resp"
    if "hello letter"                in t \
    or "welcome letter"              in t: return "NOTIFICACION_CESION",       "fn:hello/welcome"
    if "certificado correspondencia" in t: return "CERTIFICADO_CORRESPONDENCIA","fn:correspondencia"
    if "certificado saldo"           in t: return "CERTIFICADO_SALDO",         "fn:cert_saldo"
    if "certificado ic"              in t: return "CERTIFICADO_IC",            "fn:cert_ic"
    if "certificado titularidad"     in t: return "CERTIFICADO_TITULARIDAD",   "fn:cert_titularidad"

    if "movimientos" in t:
        m = re.search(r"movimientos\s+(\d+)", t)
        suf = f"_{m.group(1)}" if m else ""
        return f"MOVIMIENTOS{suf}", f"fn:movimientos{suf or '_1'}"

    if "tasa" in t:
        return "TASA", "fn:tasa"

    return "UNKNOWN", f"fn:unclassified:{base}"

# ──────────────────────────────────────────
# RESOLUCIÓN DEL NÚMERO DE DOCUMENTO EN EL ÍNDICE
# ──────────────────────────────────────────

def _first_item(items: List[IdxItem], pred) -> Optional[IdxItem]:
    """Devuelve el primer item que cumple el predicado."""
    for it in items:
        if pred(it.dn):
            return it
    return None

def _all_items(items: List[IdxItem], pred) -> List[IdxItem]:
    """Devuelve todos los items que cumplen el predicado."""
    return [it for it in items if pred(it.dn)]

def resolve_x(doc_type: str, items: List[IdxItem]) -> Optional[Tuple[int, str]]:
    """
    Devuelve (numero, descripcion_indice) para un doc_type simple
    (no CTFDOs ni MOVIMIENTOS, que se resuelven con pre-cálculo de grupos).
    """
    def _r(pred) -> Optional[Tuple[int, str]]:
        it = _first_item(items, pred)
        return (it.num, it.desc) if it else None

    if doc_type == "PODER":
        return _r(lambda d: "poder" in d)

    if doc_type == "CONTRATO":
        return _r(lambda d: "contrato" in d)

    if doc_type == "TESTIMONIO":
        return _r(lambda d: "testimonio" in d or "notarial" in d)

    if doc_type == "CERTIFICADO_TITULARIDAD":
        return _r(lambda d: "titularidad" in d)

    if doc_type == "CERTIFICADO_SALDO":
        # "saldo del cedente", "saldo deudor", etc. — siempre contienen 'saldo'
        return _r(lambda d: "saldo" in d)

    if doc_type == "CERTIFICADO_IC":
        # Standard: "Certificado de deuda"  (certificado + deuda sin 'saldo')
        # BBVA:     "Certificado deuda (Investcapital LTD)"
        # NO debe matchear: "Movimientos de origen de la deuda"
        return _r(lambda d:
            "investcapital" in d or
            ("certificado" in d and "deuda" in d and "saldo" not in d)
        )

    if doc_type == "DECLARACION_RESPONSABLE":
        return _r(lambda d: "responsab" in d)

    if doc_type == "TASA":
        return _r(lambda d:
            "tasa" in d or "autoliquidacion" in d or "modelo 696" in d
        )

    if doc_type == "NOTIFICACION_CESION":
        # Standard: "Notificación de la cesión"
        # Carrefour: "Carta de notificación de cesión del crédito"
        return _r(lambda d:
            ("notificacion" in d and "cesion" in d) or
            ("carta" in d and "notificacion" in d) or
            ("comunicacion" in d and "cesion" in d)
        )

    if doc_type == "CERTIFICADO_CORRESPONDENCIA":
        return _r(lambda d: "correspondencia" in d)

    return None

# ──────────────────────────────────────────
# BÚSQUEDA EN ÍNDICE
# ──────────────────────────────────────────

_STOP_WORDS = {
    "para", "como", "este", "esta", "pero", "todo", "cada", "bien", "otro",
    "otra", "cual", "cuyo", "hace", "algo", "tipo", "caso", "debe", "haya",
    "sean", "sera", "cabe", "dado", "dicho", "contenga", "incluya", "tenga",
    "que", "con", "del", "los", "las", "una", "por", "sin", "sus", "son",
    "hay", "mas", "muy", "cuya", "dicha",
}

def _find_in_index(idx_kw: str, items: List[IdxItem]) -> Optional[IdxItem]:
    """
    Busca en items el primero cuya descripción coincide con idx_kw.
    Extrae palabras clave (>= 4 chars, no stop words) del texto de col B.
    Usa \b...\b para evitar falsos positivos con formas en plural
    (ej. 'justificante' no debe coincidir con 'justificantes').
    Estrategia: todos los keywords → mitad → el más largo.
    """
    words_raw = al(idx_kw).split()
    keywords = []
    for w in words_raw:
        w_clean = re.sub(r'[^a-z0-9]', '', w)
        if len(w_clean) >= 4 and w_clean not in _STOP_WORDS:
            keywords.append(w_clean)
    if not keywords:
        return None

    def _hit(dn: str, kw: str) -> bool:
        return bool(re.search(rf'\b{re.escape(kw)}\b', dn))

    for item in items:
        if all(_hit(item.dn, kw) for kw in keywords):
            return item
    min_match = max(1, round(len(keywords) * 0.7))
    for item in items:
        if sum(1 for kw in keywords if _hit(item.dn, kw)) >= min_match:
            return item
    longest = max(keywords, key=len)
    for item in items:
        if _hit(item.dn, longest):
            return item
    return None

def _fn_matches(fn_kw: str, fn_norm: str) -> bool:
    """
    Comprueba que todos los tokens de fn_kw aparecen en fn_norm
    como inicio de palabra (orden independiente).
    El token 'cont' se expande a 'contenido' para evitar falsos positivos
    con 'contrato'.
    fn_norm debe tener guiones bajos sustituidos por espacios.
    """
    for tok in al(fn_kw).split():
        expanded = _TOKEN_EXPAND.get(tok, tok)
        if not re.search(rf'\b{re.escape(expanded)}', fn_norm):
            return False
    return True


def _ctfdo_type(fn_kw: str) -> Optional[Tuple[str, str]]:
    """
    Si fn_kw es una regla CTFDO devuelve (canal, tipo).
    canal: 'EMAIL' | 'SMS' | 'BUROFAX'
    tipo:  'contenido' | 'fehaciente'
    Devuelve None si no es una regla CTFDO.
    """
    t = al(fn_kw)
    for ch in ["email", "sms", "burofax"]:
        if ch in t:
            return (ch.upper(), "contenido" if "cont" in t else "fehaciente")
    return None

# ──────────────────────────────────────────
# PROCESO POR EXPEDIENTE
# ──────────────────────────────────────────

def process_exp(exp: str, exp_dir: str, rules: List[Tuple[str, str, str]],
                idx_num: Optional[int],
                asunto_codigo: str = "", common_dir: str = "",
                exp_idx: int = 0, exp_total: int = 0) -> List[Dict]:
    rows: List[Dict] = []

    # ── 1. Índice (última página de la DEMANDA) ──────────────────
    idx_pdf = extract_last_page_as_index(exp_dir, exp, exp_idx=exp_idx, exp_total=exp_total)
    items: List[IdxItem] = []
    if idx_pdf:
        items = parse_index(pdf_text(idx_pdf))
    idx_name = os.path.basename(idx_pdf) if idx_pdf else "NO_INDICE"

    all_pdfs = sorted(f for f in os.listdir(exp_dir) if f.lower().endswith(".pdf"))

    # Nombres normalizados de copias de docs comunes creadas por stamp_pdf.py
    # (formato: {exp}_{nombre_original}). Se excluyen de la clasificación para
    # evitar duplicados cuando el fichero ya está en _COMMON_DOCS.
    _common_copy_names: set = set()
    for _fc, _ in _COMMON_DOCS:
        _common_copy_names.add(al(f"{exp}_{_fc}"))
    _notarial_fn_al = al(f"{exp}_{_NOTARIAL_COMMON_DOC[0]}")
    _common_copy_names.add(_notarial_fn_al)

    # ── 2. Pre-calcular asignación posicional de CTFDOs ──────────
    cont_files_by_ch: Dict[str, List[str]] = {}
    feh_files_by_ch:  Dict[str, List[str]] = {}
    for fn in all_pdfs:
        fn_norm = al(fn)
        # Aplicar las mismas exclusiones que en el bucle principal
        if fn_norm.startswith("indice"):
            continue
        if "plantilla" in fn_norm and ("resumen" in fn_norm or fn_norm == "plantilla.pdf" or fn_norm.endswith("_plantilla.pdf")):
            continue
        if fn.lower().endswith("_ocr.pdf"):
            continue
        if "demanda" in fn_norm:
            continue
        if al(fn[:-4]) == al(f"{exp}_firmado"):
            continue
        if fn_norm in _common_copy_names:
            continue
        if "welcome" in fn_norm and "letter" in fn_norm:
            continue
        fn_match = fn_norm.replace('_', ' ')
        for fn_kw, _, _ in rules:
            if _fn_matches(fn_kw, fn_match):
                ct = _ctfdo_type(fn_kw)
                if ct:
                    ch, tipo = ct
                    (cont_files_by_ch if tipo == "contenido" else feh_files_by_ch)\
                        .setdefault(ch, []).append(fn)
                break

    contenido_items = [it for it in items if "contenido" in it.dn]
    fehaciente_items = [it for it in items if
                        ("justificante" in it.dn or
                         "intento"      in it.dn or
                         "fehaciente"   in it.dn or
                         "acuse"        in it.dn or
                         ("remision"    in it.dn and "recepcion" in it.dn))
                        and "contenido" not in it.dn]

    ctfdo_map: Dict[str, IdxItem] = {}
    for idx_list, files_by_ch in [(contenido_items, cont_files_by_ch),
                                   (fehaciente_items, feh_files_by_ch)]:
        if not idx_list:
            continue
        ordered: List[str] = []
        for ch in _CHANNEL_ORDER:
            ordered.extend(files_by_ch.get(ch, []))
        if len(idx_list) == 1:
            # Un único entry en el índice → todos los ficheros comparten ese número
            for fn in ordered:
                ctfdo_map[fn] = idx_list[0]
        else:
            # Varios entries → asignación posicional por canal
            for i, fn in enumerate(ordered):
                if i < len(idx_list):
                    ctfdo_map[fn] = idx_list[i]

    # ── Fallback contenido: cuando no hay entradas de "contenido" en el índice
    #    pero "Justificantes de intento de notificación" aparece ≥2 veces,
    #    las entradas fehaciente sobrantes (tras cubrir los ficheros fehaciente)
    #    se usan para los ficheros de contenido (burofax cont, sms cont, email cont).
    if not contenido_items:
        feh_ordered_fb = []
        for ch in _CHANNEL_ORDER:
            feh_ordered_fb.extend(feh_files_by_ch.get(ch, []))
        cont_ordered_fb = []
        for ch in _CHANNEL_ORDER:
            cont_ordered_fb.extend(cont_files_by_ch.get(ch, []))
        leftover = fehaciente_items[len(feh_ordered_fb):]   # entries fehaciente sobrantes
        if leftover and cont_ordered_fb:
            if len(leftover) == 1:
                for fn in cont_ordered_fb:
                    ctfdo_map[fn] = leftover[0]
            else:
                for i, fn in enumerate(cont_ordered_fb):
                    if i < len(leftover):
                        ctfdo_map[fn] = leftover[i]

    # ── 3. Procesar cada PDF ─────────────────────────────────────
    for fn in all_pdfs:
        fn_norm = al(fn)

        # Ficheros a ignorar
        if fn_norm.startswith("indice"):
            continue
        if fn.lower().endswith("_ocr.pdf"):
            continue
        if "demanda" in fn_norm:
            continue
        if al(fn[:-4]) == al(f"{exp}_firmado"):
            continue
        if fn_norm in _common_copy_names:
            continue
        # WELCOME LETTER → prioridad sobre el bucle de reglas para evitar
        # falsos positivos cuando el nombre contiene 'burofax', 'email', etc.
        # Búsqueda estricta: requiere 'notificacion' Y 'cesion' en la entrada
        # para evitar falsos positivos con 'Certificado notarial de la cesión'.
        if "welcome" in fn_norm and "letter" in fn_norm:
            wl_item = next(
                (it for it in items if "notificacion" in it.dn and "cesion" in it.dn),
                None
            )
            if wl_item is None:
                wl_item = next(
                    (it for it in items if "carta" in it.dn and "cesion" in it.dn),
                    None
                )
            if wl_item:
                rows.append({
                    "asunto_codigo":           asunto_codigo,
                    "referencia_demanda":      exp,
                    "nombre_fichero_original": fn,
                    "nombre_correcto":         f"DOC. {wl_item.num} {wl_item.desc}",
                    "numero_documento":        str(wl_item.num),
                    "entrada_indice":          wl_item.desc,
                    "numero_indice":           idx_num if idx_num is not None else "",
                    "status":                  "OK",
                    "motivo":                  "welcome letter",
                    "fn_kw":                   "welcome letter",
                    "ruta":                    os.path.join(exp_dir, fn),
                })
            else:
                rows.append({
                    "asunto_codigo":           asunto_codigo,
                    "referencia_demanda":      exp,
                    "nombre_fichero_original": fn,
                    "nombre_correcto":         "WELCOME LETTER",
                    "numero_documento":        "",
                    "entrada_indice":          "",
                    "numero_indice":           idx_num if idx_num is not None else "",
                    "status":                  "OK",
                    "motivo":                  "welcome_letter_fallback",
                    "fn_kw":                   "welcome letter",
                    "ruta":                    os.path.join(exp_dir, fn),
                })
            continue

        # PLANTILLAS RESUMEN → añadir directamente sin cruzar reglas ni índice
        if "plantilla" in fn_norm and ("resumen" in fn_norm or fn_norm == "plantilla.pdf" or fn_norm.endswith("_plantilla.pdf")):
            rows.append({
                "asunto_codigo":           asunto_codigo,
                "referencia_demanda":      exp,
                "nombre_fichero_original": fn,
                "nombre_correcto":         "PLANTILLAS RESUMEN",
                "numero_documento":        "",
                "entrada_indice":          "",
                "numero_indice":           idx_num if idx_num is not None else "",
                "status":                  "OK",
                "motivo":                  "plantilla_resumen",
                "ruta":                    os.path.join(exp_dir, fn),
            })
            continue

        row: Dict = {
            "asunto_codigo":           asunto_codigo,
            "referencia_demanda":      exp,
            "nombre_fichero_original": fn,
            "nombre_correcto":         "",
            "numero_documento":        "",
            "entrada_indice":          "",
            "numero_indice":           idx_num if idx_num is not None else "",
            "status":                  "",
            "motivo":                  "",
            "fn_kw":                   "",
            "ruta":                    os.path.join(exp_dir, fn),
        }

        # ── Buscar regla por nombre de fichero (col A) ──────────
        fn_match = fn_norm.replace('_', ' ')

        # ── Detección específica MOVIMIENTOS vs MOVIMIENTOS 2 ────
        # Los tokens sueltos "1"/"2" en fn_match pueden coincidir con el
        # número de carpeta (IBBVA_1, IBBVA_2), no con el documento.
        # Se busca la secuencia literal "movimientos 2" / "movimientos"
        # en el nombre del fichero para forzar la regla correcta.
        _mov_force: Optional[str] = None
        if re.search(r'\bmovimientos\s*2\b', fn_match):
            _mov_force = "movimientos 2"
        elif re.search(r'\bmovimientos\b', fn_match):
            _mov_force = "movimientos 1"

        # Recoge TODAS las reglas que coinciden con el nombre del fichero
        # (puede haber varias con el mismo fn_kw pero distinto idx_kw)
        matched_rules: List[Tuple[str, str, str]] = []

        if _mov_force:
            # Para ficheros de movimientos, seleccionar las reglas por fn_kw
            # exacto, sin pasar por _fn_matches (que confunde el número de la
            # carpeta IBBVA_N con el número del documento).
            matched_rules = [(fk, ik, tmpl) for fk, ik, tmpl in rules
                             if al(fk) == _mov_force]
            if not matched_rules:
                # Fallback: usar la regla genérica MOVIMIENTOS
                matched_rules = [(fk, ik, tmpl) for fk, ik, tmpl in rules
                                 if al(fk) == "movimientos"]
        else:
            seen_fn_kw: Optional[str] = None
            for fn_kw, idx_kw, template in rules:
                if _fn_matches(fn_kw, fn_match):
                    if seen_fn_kw is None:
                        seen_fn_kw = fn_kw
                    if fn_kw == seen_fn_kw:
                        matched_rules.append((fn_kw, idx_kw, template))
                    else:
                        break  # fn_kw distinto → regla de menor prioridad, parar

        if not matched_rules:
            row["status"] = "FAIL"
            row["motivo"] = f"No clasificado: {fn}"
            rows.append(row)
            continue

        fn_kw = matched_rules[0][0]

        if not items:
            row["status"] = "FAIL"
            row["motivo"] = f"Sin índice ({idx_name})"
            rows.append(row)
            continue

        # ── Resolver número de documento ─────────────────────────
        ct = _ctfdo_type(fn_kw)
        if ct:
            item = ctfdo_map.get(fn)
            if item is None:
                row["status"] = "FAIL"
                row["motivo"] = f"CTFDO sin número en índice para '{fn_kw}' | idx:{idx_name}"
                rows.append(row)
                continue
            idx_kw   = matched_rules[0][1]
            template = matched_rules[0][2]
        else:
            # Probar cada idx_kw de las reglas coincidentes hasta encontrar uno en el índice
            item     = None
            idx_kw   = matched_rules[0][1]
            template = matched_rules[0][2]
            for _fk, _ik, _tmpl in matched_rules:
                _item = _find_in_index(_ik, items)
                if _item is not None:
                    item, idx_kw, template = _item, _ik, _tmpl
                    break
            if item is None and "hello" in al(fn_kw):
                item = _find_in_index("Carta notificacion cesion credito", items)
            if item is None and "testimonio" in al(fn_kw):
                item = _find_in_index("Certificado notarial de cesion de credito", items)
            if item is None and "movimiento" in al(fn_kw) and "2" not in fn_kw:
                item = _find_in_index("Detalle de los cargos", items)
            if item is None and "movimientos 2" in al(fn_kw):
                item = _find_in_index("Liquidacion tras el pase a mora", items)
            if item is None:
                row["status"] = "FAIL"
                row["motivo"] = f"Sin número en índice para '{idx_kw}' | idx:{idx_name}"
                rows.append(row)
                continue

        # ── Construir nombre final ────────────────────────────────
        row["nombre_correcto"]  = f"DOC. {item.num} {item.desc}"
        row["numero_documento"] = str(item.num)
        row["entrada_indice"]   = item.desc
        row["status"]           = "OK"
        row["motivo"]           = fn_kw
        row["fn_kw"]            = fn_kw
        rows.append(row)

    # ── Documentos comunes desde ROOT ────────────────────────────
    if common_dir and items:
        for fn_common, idx_kw in _COMMON_DOCS:
            item = _find_in_index(idx_kw, items)
            # PODER: si no aparece en el índice, forzar como DOC. 1
            is_poder = "poder" in al(idx_kw)
            if item is None and is_poder:
                nombre_correcto  = "DOC. 1 Poder general para pleitos"
                numero_documento = "1"
                entrada_indice   = "Poder general para pleitos"
                status           = "OK"
                motivo           = f"{idx_kw} (forzado DOC. 1 por ausencia en índice)"
            else:
                nombre_correcto  = f"DOC. {item.num} {item.desc}" if item else ""
                numero_documento = str(item.num) if item else ""
                entrada_indice   = item.desc if item else ""
                status           = "OK" if item else "FAIL"
                motivo           = idx_kw if item else f"Sin número en índice para '{idx_kw}'"
            rows.append({
                "asunto_codigo":           asunto_codigo,
                "referencia_demanda":      exp,
                "nombre_fichero_original": fn_common,
                "nombre_correcto":         nombre_correcto,
                "numero_documento":        numero_documento,
                "entrada_indice":          entrada_indice,
                "numero_indice":           idx_num if idx_num is not None else "",
                "status":                  status,
                "motivo":                  motivo,
                "fn_kw":                   idx_kw,
                "ruta":                    os.path.join(common_dir, fn_common),
            })

    # ── Fila del fichero de índice ────────────────────────────────
    if idx_pdf:
        rows.append({
            "asunto_codigo":           asunto_codigo,
            "referencia_demanda":      exp,
            "nombre_fichero_original": os.path.basename(idx_pdf),
            "nombre_correcto":         "Indice",
            "numero_documento":        "",
            "entrada_indice":          "",
            "numero_indice":           idx_num if idx_num is not None else "",
            "status":                  "OK",
            "motivo":                  "indice",
            "fn_kw":                   "INDICE",
            "ruta":                    idx_pdf,
        })

    # ── Detectar duplicados de numero_documento ───────────────────
    from collections import Counter
    num_counts = Counter(
        r["numero_documento"] for r in rows
        if r["status"] == "OK" and r["numero_documento"]
    )
    for r in rows:
        if r["status"] == "OK" and num_counts.get(r["numero_documento"], 0) > 1:
            r["status"] = "FAIL"
            r["motivo"] = f"Número de documento duplicado: {r['numero_documento']} | {r['motivo']}"

    return rows

# ──────────────────────────────────────────
# DEMANDAS CSV
# ──────────────────────────────────────────

def _macro_from_tipo(tipo: str) -> str:
    """Devuelve el código MACRO según el tipo de procedimiento."""
    t = al(tipo)
    if "verbal"                      in t: return "233"
    if "monitorio"                   in t: return "383"
    if "hipotecario"                 in t: return "343"
    if "ordinario"                   in t: return "246"
    if "ejecucion de titulo judicial" in t or "etj" in t: return "262"
    if "ejecucion de titulo no judicial" in t or "etnj" in t or "enj" in t: return "299"
    return ""


def collect_demandas(in_root: str, exps: List[str],
                     codes_map: Dict[str, str],
                     datatape: Dict[str, Dict],
                     folder_alias: Dict[str, str] = None) -> List[Dict]:
    """
    Recorre todas las carpetas de expedientes y recoge los PDFs cuyo nombre
    contiene la palabra 'demanda'.
    Devuelve filas para demandas.xlsx.
    """
    if folder_alias is None:
        folder_alias = {}
    rows: List[Dict] = []
    for exp_folder in exps:
        exp_key = exp_folder if exp_folder in datatape else folder_alias.get(exp_folder, exp_folder)
        exp_dir = os.path.join(in_root, exp_folder)
        info    = datatape.get(exp_key, {})
        tipo    = info.get("tipo_proc", "")
        city    = info.get("city", "")
        macro   = _macro_from_tipo(tipo)
        # Clase: DEMANDA MON si es monitorio de plaza especial
        is_mon_especial = (
            "monitorio" in al(tipo) and
            al(city) in _MONITORIO_EXCLUSION_CITIES_NORM
        )
        clase_demanda = "DEMANDA MON" if is_mon_especial else "DEMANDA"
        pdfs    = sorted(f for f in os.listdir(exp_dir) if f.lower().endswith(".pdf"))

        # Buscar PDFs con "demanda" en el nombre
        demanda_found = False
        for fn in pdfs:
            if "demanda" in al(fn):
                demanda_found = True
                rows.append({
                    "asunto_codigo":           codes_map.get(exp_key, ""),
                    "referencia_demanda":      exp_folder,
                    "nombre_fichero_original": fn,
                    "CLASE":                   clase_demanda,
                    "texto":                   "DEMANDA",
                    "ruta":                    os.path.join(exp_dir, fn),
                    "MACRO":                   macro,
                })

        # Fallback: si no hay "demanda", buscar fichero <expediente>_firmado.pdf
        if not demanda_found:
            firmado_name = f"{exp_folder}_firmado.pdf"
            for fn in pdfs:
                if fn.lower() == firmado_name.lower():
                    print(f"[demandas] WARN '{exp_folder}': sin DEMANDA, usando fallback '{fn}'")
                    rows.append({
                        "asunto_codigo":           codes_map.get(exp_key, ""),
                        "referencia_demanda":      exp_folder,
                        "nombre_fichero_original": fn,
                        "CLASE":                   clase_demanda,
                        "texto":                   "DEMANDA",
                        "ruta":                    os.path.join(exp_dir, fn),
                        "MACRO":                   macro,
                    })
                    break
    return rows


# ──────────────────────────────────────────
# VALIDACIÓN DE COBERTURA DE ÍNDICES
# ──────────────────────────────────────────

def _idx_kw_covers_strict(idx_kw: str, item: IdxItem) -> bool:
    """
    Versión estricta para validación: requiere que TODOS los keywords
    del idx_kw aparezcan en la descripción del item (sin umbral del 70%).
    Evita falsos positivos cuando se prueba con un único item.
    """
    words_raw = al(idx_kw).split()
    keywords = []
    for w in words_raw:
        w_clean = re.sub(r'[^a-z0-9]', '', w)
        if len(w_clean) >= 4 and w_clean not in _STOP_WORDS:
            keywords.append(w_clean)
    if not keywords:
        return False
    # s? al final permite tolerar diferencias singular/plural (ej. "justificante" ↔ "justificantes")
    return all(bool(re.search(rf'\b{re.escape(kw)}s?\b', item.dn)) for kw in keywords)

# Patrones de índice cubiertos por lógica especial o aliases que el proceso resuelve
# correctamente pero el validador estricto no llega a cubrir por el idx_kw exacto.
_CTFDO_IDX_PATTERNS: List[str] = [
    "Certificado de remision recepcion",          # acuse de recibo / envío certificado
    "Declaracion responsable",                    # cubierto por regla "decl responsable" → CERTIFICADO_1
    "Certificado notarial de cesion de credito",  # texto antiguo de TESTIMONIO (fallback)
    "Detalle de los cargos",                      # texto alternativo de MOVIMIENTOS
]

def validate_index_coverage(indices_dir: str,
                             rules: List[Tuple[str, str, str]],
                             root: str,
                             common_docs: List[Tuple[str, str]] = None,
                             used_idx_nums: set = None) -> None:
    """
    Lee los PDFs de índices usados por el datatape, deduplica las entradas
    por texto (ignorando el número de documento) y comprueba cuáles están
    cubiertas por las reglas actuales (idx_kw).

    - Omite y registra en log los índices no referenciados por ningún expediente.
    - Imprime en log las entradas SIN regla asignada.
    - Genera reglas_indices_remesa.csv con todas las entradas y su estado.
    """
    if not os.path.isdir(indices_dir):
        print("[validación] Carpeta de índices no encontrada, se omite la validación.")
        return

    # ── 1. Recopilar entradas únicas de los índices usados ────────
    # key: al(desc)  →  {desc: str original, indices: set de números}
    all_entries: Dict[str, Dict] = {}
    n_indices      = 0
    skipped_files  = []

    for fn in sorted(os.listdir(indices_dir)):
        if not fn.lower().endswith(".pdf"):
            continue
        m = re.match(r"indice\s+(\d+)", al(fn))
        idx_num_str = m.group(1) if m else None

        # Saltar índices no referenciados por ningún expediente del datatape
        if used_idx_nums is not None and idx_num_str is not None:
            if int(idx_num_str) not in used_idx_nums:
                skipped_files.append(fn)
                continue

        n_indices += 1
        items = parse_index(pdf_text(os.path.join(indices_dir, fn)))
        for item in items:
            if item.dn not in all_entries:
                all_entries[item.dn] = {"desc": item.desc, "indices": set()}
            all_entries[item.dn]["indices"].add(idx_num_str or "?")

    if skipped_files:
        print(f"[validación índices] {len(skipped_files)} índice(s) no usados en esta remesa (omitidos):")
        for fn in skipped_files:
            print(f"  –  {fn}")

    if not all_entries:
        print("[validación] No se encontraron entradas en los índices.")
        return

    # ── 2. Para cada entrada, buscar qué regla la cubre ──────────
    # Usamos un item temporal con num=0 para probar cada regla.
    results = []
    unmapped: List[str] = []

    for dn in sorted(all_entries):
        info = all_entries[dn]
        desc = info["desc"]
        indices_sorted = sorted(info["indices"],
                                key=lambda x: int(x) if x.isdigit() else 9999)
        indices_str = ", ".join(indices_sorted)

        test_item = IdxItem(0, desc, dn)
        matched_fn_kw   = ""
        matched_nombre  = ""

        for fn_kw, idx_kw, template in rules:
            if _idx_kw_covers_strict(idx_kw, test_item):
                matched_fn_kw  = fn_kw
                matched_nombre = re.sub(r"DOC\.\s*X\s*", "", template).strip()
                break

        # Si no lo cubre ninguna regla, comprobar documentos comunes (doc_comun/)
        if not matched_fn_kw and common_docs:
            for fn_common, idx_kw_common in common_docs:
                if _idx_kw_covers_strict(idx_kw_common, test_item):
                    matched_fn_kw  = "doc_comun"
                    matched_nombre = fn_common.replace(".pdf", "")
                    break

        # Si tampoco, comprobar patrones cubiertos por lógica CTFDO especial
        if not matched_fn_kw:
            for idx_kw_ctfdo in _CTFDO_IDX_PATTERNS:
                if _idx_kw_covers_strict(idx_kw_ctfdo, test_item):
                    matched_fn_kw  = "ctfdo"
                    matched_nombre = "CERTIFICADO FEHACIENTE (lógica CTFDO)"
                    break

        mapeado = "Sí" if matched_fn_kw else "No"
        if not matched_fn_kw:
            unmapped.append(f"[índices {indices_str}] {desc}")

        results.append({
            "texto_indice":    desc,
            "indices":         indices_str,
            "mapeado":         mapeado,
            "fn_kw":           matched_fn_kw,
            "nombre_resultado": matched_nombre,
        })

    # ── 3. Log ───────────────────────────────────────────────────
    sep = "=" * 60
    print(f"\n{sep}")
    print(f"[validación índices] {n_indices} índices | "
          f"{len(all_entries)} entradas únicas")
    print(f"[validación índices] Con regla : {len(all_entries) - len(unmapped)}")
    print(f"[validación índices] SIN regla : {len(unmapped)}")
    if unmapped:
        print("[validación índices] ⚠ Entradas SIN regla asignada:")
        for u in unmapped:
            print(f"  ⚠  {u}")
    print(f"{sep}\n")

    # ── 4. CSV ───────────────────────────────────────────────────
    csv_path = os.path.join(root, "reglas_indices_remesa.csv")
    fields = ["texto_indice", "indices", "mapeado", "fn_kw", "nombre_resultado"]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        w.writeheader()
        w.writerows(results)
    print(f"[validación índices] CSV generado: {csv_path}")


# ──────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────

def main(root: str) -> None:
    in_root = os.path.join(root, "IN")

    if not os.path.isdir(in_root):
        raise FileNotFoundError(f"No existe la carpeta IN/: {in_root}")

    rules = _RULES

    datatape_path = os.path.join(in_root, "datatape.xlsx")
    if not os.path.isfile(datatape_path):
        raise FileNotFoundError(f"No se encontró el datatape: {datatape_path}")
    exp_idx_map = load_datatape(datatape_path)

    exps = sorted(
        d for d in os.listdir(in_root)
        if os.path.isdir(os.path.join(in_root, d)) and al(d) not in ("indices", "iindices")
    )

    # ── Códigos Kmaleon ────────────────────────────────────────────
    codes_map = load_kmaleon_codes(exp_idx_map)

    # Fallback: si la carpeta no coincide con Original Contract Number,
    # buscar por Whole Case Number con "/" → "_"
    folder_alias: Dict[str, str] = {
        info["referencia2"].replace("/", "_"): exp
        for exp, info in exp_idx_map.items()
        if "/" in info.get("referencia2", "")
    }

    all_rows: List[Dict] = []
    total_exps = len(exps)
    for i, exp_folder in enumerate(exps, 1):
        exp_key = exp_folder if exp_folder in exp_idx_map else folder_alias.get(exp_folder, exp_folder)
        info    = exp_idx_map.get(exp_key, {})
        idx_num = info.get("idx_num") if info else None
        if idx_num is None:
            print(f"[WARN] Expediente '{exp_folder}' no encontrado en datatape.xlsx")
        exp_rows = process_exp(
            exp_folder, os.path.join(in_root, exp_folder), rules, idx_num,
            asunto_codigo=codes_map.get(exp_key, ""),
            common_dir=os.path.join(root, "doc_comun"),
            exp_idx=i, exp_total=total_exps,
        )
        all_rows.extend(exp_rows)

    out_path = os.path.join(root, "documentos.xlsx")
    fields = ["asunto_codigo", "codigo_2", "ruta", "Texto", "Fecha", "Clase", "status"]
    xlsx_rows = []
    for r in all_rows:
        exp      = r["referencia_demanda"]
        exp_key  = exp if exp in exp_idx_map else folder_alias.get(exp, exp)
        info     = exp_idx_map.get(exp_key, {})
        tipo_proc = info.get("tipo_proc", "")
        city      = info.get("city", "")
        is_plantilla = r.get("nombre_correcto") == "PLANTILLAS RESUMEN"

        # Si el documento estaba OK pero no tiene código kmaleon → FAIL
        status = r["status"]
        if status == "OK" and not r.get("asunto_codigo"):
            status = f"FAIL: Expediente {exp} no encontrado en kmaleon"

        clase = (
            ""
            if status != "OK" or is_plantilla
            else _get_clase(r.get("fn_kw", ""), tipo_proc, city)
        )
        xlsx_rows.append({
            "asunto_codigo": r["asunto_codigo"],
            "codigo_2":      "",
            "ruta":          r["ruta"],
            "Texto":         r["nombre_correcto"],
            "Fecha":         "",
            "Clase":         clase,
            "status":        status,
        })
    import openpyxl as _openpyxl
    wb_out = _openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append(fields)
    for row in xlsx_rows:
        ws_out.append([row[f] for f in fields])
    wb_out.save(out_path)

    ok   = sum(1 for r in all_rows if r["status"] == "OK")
    fail = sum(1 for r in all_rows if r["status"] == "FAIL")
    print(f"OK: {ok}  |  FAIL: {fail}  |  Total filas: {len(all_rows)}")
    print(f"Excel generado: {out_path}")

    # ── demandas.xlsx ──────────────────────────────────────────────
    demandas_rows = collect_demandas(in_root, exps, codes_map, exp_idx_map, folder_alias)

    # Añadir TESTIMONIO FUSION solo a expedientes cuyo índice tiene "Escritura de fusión"
    _notarial_fn, _notarial_idx_kw = _NOTARIAL_COMMON_DOC
    _notarial_path = os.path.join(root, "doc_comun", _notarial_fn)
    # Construir set de expedientes que tienen la entrada en el índice (status OK)
    _exps_with_escritura: set = {
        r["referencia_demanda"]
        for r in all_rows
        if r.get("status") == "OK"
        and al(_notarial_idx_kw) in al(r.get("entrada_indice", ""))
    }
    for exp_folder in sorted(_exps_with_escritura):
        exp_key   = exp_folder if exp_folder in exp_idx_map else folder_alias.get(exp_folder, exp_folder)
        info      = exp_idx_map.get(exp_key, {})
        tipo_proc = info.get("tipo_proc", "")
        city      = info.get("city", "")
        is_mon_especial = (
            "monitorio" in al(tipo_proc)
            and al(city) in _MONITORIO_EXCLUSION_CITIES_NORM
        )
        clase = "NOTARIAL MON" if is_mon_especial else "NOTARIAL"
        demandas_rows.append({
            "asunto_codigo":           codes_map.get(exp_key, ""),
            "referencia_demanda":      exp_folder,
            "nombre_fichero_original": _notarial_fn,
            "CLASE":                   clase,
            "texto":                   "NOTARIAL",
            "ruta":                    _notarial_path,
            "MACRO":                   "",
        })
    if _exps_with_escritura:
        print(f"[demandas] NOTARIAL añadido a {len(_exps_with_escritura)} expediente(s) con 'Escritura de fusión' en el índice")
    demandas_path   = os.path.join(root, "demandas.xlsx")
    demandas_fields = ["asunto_codigo", "codigo_2", "ruta", "texto", "Fecha", "Clase", "Opciones", "Macro"]
    wb_d = _openpyxl.Workbook()
    ws_d = wb_d.active
    ws_d.append(demandas_fields)
    for r in demandas_rows:
        ws_d.append([
            r["asunto_codigo"], "", r["ruta"], r["texto"], "",
            r["CLASE"], "", r["MACRO"],
        ])
    wb_d.save(demandas_path)
    print(f"Demandas      : {len(demandas_rows)} ficheros → {demandas_path}")

    # Conteo de macros
    from collections import Counter
    macro_counts = Counter(r["MACRO"] or "Sin macro" for r in demandas_rows)
    macro_labels = {
        "233": "Verbal",
        "383": "Monitorio",
        "343": "Hipotecario",
        "246": "Ordinario",
        "262": "ETJ",
        "299": "ETNJ/ENJ",
        "Sin macro": "Sin macro",
    }
    print("[demandas] Distribución por MACRO:")
    for macro, count in sorted(macro_counts.items()):
        label = macro_labels.get(macro, macro)
        print(f"  {macro:>9}  ({label:<12})  →  {count} demandas")


# ──────────────────────────────────────────
# CONVERSIÓN DE RUTAS PARA WINDOWS
# ──────────────────────────────────────────

def _to_windows_path(ruta: str) -> str:
    """
    Convierte rutas Linux del fileserver a formato Windows con unidad mapeada.
    /fileserver05/SFTP/KRUK/... → K:\\SFTP\\KRUK\\...
    """
    ruta = ruta.replace("/fileserver05", "K:")
    ruta = ruta.replace("/", "\\")
    return ruta


if __name__ == "__main__":
    root_dir = sys.argv[1] if len(sys.argv) > 1 else "/fileserver05/SFTP/KRUK/DEMANDAS"
    main(root_dir)
