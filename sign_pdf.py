#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sign_pdf.py — KRUK / FIRMAR DEMANDAS

Lee demandas.csv y firma electrónicamente cada PDF con el certificado
instalado en el Keychain de macOS (o desde un fichero .p12 si se prefiere).

La firma es VISIBLE y aparece en la esquina inferior derecha de la
última página de cada PDF.

Los PDFs firmados se guardan en OUT_Signed/ con la misma estructura
de subcarpetas que IN/.

Uso:
  python3 sign_pdf.py [ROOT_DIR] [CERT_P12_opcional]

  ROOT_DIR        Carpeta raíz DEMANDAS (default: /Users/sergio.garcia/Desktop/DEMANDAS)
  CERT_P12        Ruta a un fichero .p12/.pfx (opcional).
                  Si se omite, se exporta automáticamente desde el Keychain de macOS.

La contraseña se pide siempre por teclado (no queda en el historial del shell).
  · Con Keychain: es la contraseña que protegerá el fichero temporal de exportación.
    macOS mostrará además un diálogo de autorización de acceso a la clave privada.
  · Con .p12:     es la contraseña del fichero .p12.
"""

import os
import csv
import sys
from typing import Dict, List, Optional

# ──────────────────────────────────────────
# VENV AUTO-ACTIVACIÓN
# ──────────────────────────────────────────

def _ensure_venv() -> None:
    import platform as _platform
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    if _platform.system() == "Windows":
        _venv_python = os.path.join(_script_dir, "venv", "Scripts", "python.exe")
    else:
        _venv_python = os.path.join(_script_dir, "venv", "bin", "python3")
    if os.path.isfile(_venv_python):
        if os.path.realpath(sys.executable) != os.path.realpath(_venv_python):
            print(f"[venv] Re-ejecutando con: {_venv_python}")
            if _platform.system() == "Windows":
                import subprocess as _sp
                sys.exit(_sp.run([_venv_python] + sys.argv).returncode)
            else:
                os.execv(_venv_python, [_venv_python] + sys.argv)

_ensure_venv()

# ──────────────────────────────────────────
# AUTO-INSTALACIÓN DE DEPENDENCIAS
# ──────────────────────────────────────────

def _ensure_deps() -> None:
    _deps = [
        ("pypdf",                    "pypdf"),
        ("pyhanko",                  "pyhanko"),
        ("pyhanko_certvalidator",    "pyhanko-certvalidator"),
    ]
    import importlib.util, subprocess
    for module, package in _deps:
        if importlib.util.find_spec(module) is None:
            print(f"[deps] Instalando {package} …")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])

_ensure_deps()

# ──────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────

SIGNATURE_FIELD_NAME = "FirmaElectronica"
SIGNATURE_REASON     = "Firma del documento de demanda"
SIGNATURE_LOCATION   = "España"
SIGNATURE_CONTACT    = ""   # email del firmante (opcional)

# Ruta al certificado .p12 (en el fileserver, accesible desde la máquina Jenkins)
CERT_P12_PATH = "/fileserver05/SFTP/KRUK/DEMANDAS/doc_comun/Certificados.p12"

# Tamaño y posición del recuadro de firma (en puntos PDF, 1 pt = 1/72 pulgada)
SIG_BOX_WIDTH  = 210   # ≈ 7.4 cm
SIG_BOX_HEIGHT = 65    # ≈ 2.3 cm
SIG_MARGIN_R   = 20    # margen desde el borde derecho
SIG_MARGIN_B   = 20    # margen desde el borde inferior

# ──────────────────────────────────────────
# CARGA DEL CERTIFICADO
# ──────────────────────────────────────────

def _load_from_p12(p12_path: str, passphrase: bytes):
    """Carga el firmante desde un fichero .p12 / .pfx."""
    from pyhanko.sign import signers
    return signers.SimpleSigner.load_pkcs12(
        pfx_file=p12_path,
        passphrase=passphrase,
    )


def load_signer(cert_arg: str, passphrase: bytes):
    """Carga el firmante desde un fichero .p12 / .pfx."""
    if not cert_arg:
        raise ValueError(
            "No se ha indicado ruta al certificado .p12.\n"
            "Configura CERT_P12_PATH en el script o pásala como segundo argumento."
        )
    if not os.path.isfile(cert_arg):
        raise FileNotFoundError(
            f"\nNo se encontró el fichero de certificado: {cert_arg}\n"
            "Comprueba que el fichero Certificados.p12 está en la carpeta DEMANDAS."
        )
    print(f"[cert] Cargando desde fichero: {os.path.basename(cert_arg)}")
    signer = _load_from_p12(cert_arg, passphrase)
    print("[cert] Certificado cargado.")
    return signer

# ──────────────────────────────────────────
# ESTILO VISUAL DE LA FIRMA
# ──────────────────────────────────────────

def _make_stamp_style():
    """Firma visible sin recuadro, con nombre del firmante y fecha/hora."""
    from pyhanko.stamp import TextStampStyle
    return TextStampStyle(
        stamp_text="Firmado digitalmente\n%(signer)s\n%(ts)s",
        border_width=0,
    )

# ──────────────────────────────────────────
# FIRMA DE UN PDF
# ──────────────────────────────────────────

def _last_page_info(src: str):
    """Devuelve (índice_última_página, ancho_pts, alto_pts)."""
    from pypdf import PdfReader
    reader = PdfReader(src)
    n      = len(reader.pages)
    page   = reader.pages[n - 1]
    return n - 1, float(page.mediabox.width), float(page.mediabox.height)


def sign_pdf(src: str, dst: str, signer, stamp_style) -> None:
    """
    Firma digitalmente src y guarda en dst.
    La firma visible se sitúa en la esquina inferior derecha de la última página.
    """
    from pyhanko.sign.signers.pdf_signer import PdfSigner, PdfSignatureMetadata
    from pyhanko.sign.fields import SigFieldSpec
    from pyhanko.pdf_utils.incremental_writer import IncrementalPdfFileWriter

    last_idx, page_w, page_h = _last_page_info(src)

    # Coordenadas del recuadro (origen PDF = esquina inferior izquierda)
    x1 = page_w - SIG_BOX_WIDTH  - SIG_MARGIN_R
    y1 = SIG_MARGIN_B
    x2 = page_w - SIG_MARGIN_R
    y2 = SIG_MARGIN_B + SIG_BOX_HEIGHT

    dst_dir = os.path.dirname(dst)
    if dst_dir:
        os.makedirs(dst_dir, exist_ok=True)

    meta = PdfSignatureMetadata(
        field_name=SIGNATURE_FIELD_NAME,
        reason=SIGNATURE_REASON,
        location=SIGNATURE_LOCATION,
        contact_info=SIGNATURE_CONTACT or None,
    )

    field_spec = SigFieldSpec(
        SIGNATURE_FIELD_NAME,
        on_page=last_idx,
        box=(x1, y1, x2, y2),
    )

    pdf_signer = PdfSigner(
        meta,
        signer=signer,
        stamp_style=stamp_style,
        new_field_spec=field_spec,
    )

    with open(src, "rb") as inf:
        w = IncrementalPdfFileWriter(inf)
        with open(dst, "wb") as outf:
            pdf_signer.sign_pdf(w, output=outf)

# ──────────────────────────────────────────
# LECTURA DE demandas.csv
# ──────────────────────────────────────────

def load_demandas(path: str) -> List[Dict]:
    import openpyxl as _openpyxl
    rows: List[Dict] = []
    wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(values_only=True))]
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(raw)}
        if row.get("ruta", "").strip():
            rows.append(row)
    wb.close()
    return rows

# ──────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────

def _all_signed(log_path: str) -> bool:
    """
    Devuelve True si sign_log.csv existe y TODOS los documentos tienen status OK.
    Devuelve False si no existe, está vacío o hay algún FAIL.
    """
    if not os.path.isfile(log_path):
        return False
    try:
        with open(log_path, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f, delimiter=";"))
        if not rows:
            return False
        all_ok = all((r.get("status") or "").strip().upper() == "OK" for r in rows)
        if all_ok:
            print(f"[sign] sign_log.csv existe y todos los documentos están firmados OK "
                  f"({len(rows)} ficheros). Se omite el proceso de firma.")
        return all_ok
    except Exception as e:
        print(f"[sign] No se pudo leer sign_log.csv: {e}")
        return False


def main(root: str, cert_arg: str, passphrase: bytes) -> None:
    demandas_csv = os.path.join(root, "demandas.xlsx")
    log_path     = os.path.join(root, "sign_log.csv")

    if not os.path.isfile(demandas_csv):
        raise FileNotFoundError(f"No se encuentra demandas.csv: {demandas_csv}")

    # ── Comprobar si ya están todos firmados ──────────────────────
    if _all_signed(log_path):
        return

    # Cargar certificado y estilo de firma una sola vez
    signer      = load_signer(cert_arg, passphrase)
    stamp_style = _make_stamp_style()

    demandas = load_demandas(demandas_csv)
    print(f"\n[demandas] {len(demandas)} ficheros a firmar (en su misma ruta)\n")

    log_rows: List[Dict] = []

    for row in demandas:
        src = (row.get("ruta") or "").strip()
        fn  = os.path.basename(src)

        log_row: Dict = {
            "asunto_codigo": (row.get("asunto_codigo") or "").strip(),
            "fichero":       fn,
            "ruta":          src,
            "status":        "",
            "motivo":        "",
        }

        if not os.path.isfile(src):
            log_row["status"] = "FAIL"
            log_row["motivo"] = f"Fichero no encontrado: {src}"
            log_rows.append(log_row)
            print(f"  ✗  {fn}  →  FAIL (no encontrado)")
            continue

        # Firmar a fichero temporal y sustituir el original al terminar
        tmp_path = src + ".signing_tmp"
        try:
            sign_pdf(src, tmp_path, signer, stamp_style)
            os.replace(tmp_path, src)   # sustitución atómica
            log_row["status"] = "OK"
            log_row["motivo"] = "Firmado en su ruta original"
            print(f"  ✓  {fn}")
        except Exception as e:
            # Limpiar el temporal si quedó a medias
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
            log_row["status"] = "FAIL"
            log_row["motivo"] = f"{type(e).__name__}: {e}"
            print(f"  ✗  {fn}  →  {type(e).__name__}: {e}")

        log_rows.append(log_row)

    # ── Log ──────────────────────────────────────────────────────
    log_path   = os.path.join(root, "sign_log.csv")
    log_fields = ["asunto_codigo", "fichero", "ruta", "status", "motivo"]
    with open(log_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=log_fields, delimiter=";")
        w.writeheader()
        w.writerows(log_rows)

    ok   = sum(1 for r in log_rows if r["status"] == "OK")
    fail = sum(1 for r in log_rows if r["status"] == "FAIL")
    print(f"\nFirmados OK : {ok}")
    print(f"Errores     : {fail}")
    print(f"Log         : {log_path}")


if __name__ == "__main__":
    root_dir   = sys.argv[1] if len(sys.argv) > 1 else "/fileserver05/SFTP/KRUK/DEMANDAS"
    cert_arg   = sys.argv[2] if len(sys.argv) > 2 else CERT_P12_PATH
    passphrase = b"Pintor31!"

    main(root_dir, cert_arg, passphrase)
