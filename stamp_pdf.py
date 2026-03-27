#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
stamp_pdf.py — KRUK / PLAZAS ESPECIALES

Lee los expedientes a procesar desde datatape.xlsx (filtrando por City Court)
y el mapeo de rutas/nombres desde naming_output.csv.
Para cada expediente de las plazas indicadas:
  - Copia los PDFs con status=OK a OUT_Stamped/<expediente>/
  - Estampa la primera página con el texto de la columna 'Texto'
  - Excluye documentos DEMANDA y PLANTILLAS RESUMEN
  - Genera stamp_log.csv con el resultado

Uso:
  python3 stamp_pdf.py [ROOT_DIR]

Defaults:
  ROOT_DIR     = /Users/sergio.garcia/Desktop/DEMANDAS
  DATATAPE     = ROOT_DIR/IN/datatape.xlsx
  NAMING_CSV   = ROOT_DIR/naming_output.csv
"""

import os
import re
import csv
import sys
import unicodedata
from io import BytesIO
from typing import Dict, List, Set, Tuple

# ──────────────────────────────────────────
# VENV AUTO-ACTIVACIÓN
# ──────────────────────────────────────────

def _ensure_venv() -> None:
    import platform as _platform
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    if _platform.system() == "Windows":
        _venv_python = os.path.join(_script_dir, "venv", "Scripts", "python.exe")
    else:
        _venv_python = os.path.join(_script_dir, "venv", "bin", "python3")
    if os.path.isfile(_venv_python):
        if os.path.realpath(sys.executable) != os.path.realpath(_venv_python):
            print(f"[venv] Re-ejecutando con: {_venv_python}")
            if _platform.system() == "Windows":
                import subprocess as _sp
                sys.exit(_sp.run([_venv_python] + sys.argv).returncode)
            else:
                os.execv(_venv_python, [_venv_python] + sys.argv)

_ensure_venv()

# ──────────────────────────────────────────
# AUTO-INSTALACIÓN DE DEPENDENCIAS
# ──────────────────────────────────────────

def _ensure_deps() -> None:
    _deps = [
        ("pypdf",     "pypdf"),
        ("reportlab", "reportlab"),
        ("PIL",       "pillow"),
        ("openpyxl",  "openpyxl"),
    ]
    import importlib.util, subprocess
    for module, package in _deps:
        if importlib.util.find_spec(module) is None:
            print(f"[deps] Instalando {package} …")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])

_ensure_deps()

# ──────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────

MARGIN_TOP_MM  = 3.0
MARGIN_LEFT_MM = 3.0
MAX_RIGHT_MM   = 3.0
FONT_NAME      = "Helvetica-Bold"
FONT_SIZE      = 14
BOX_PADDING_PT = 4.0
BOX_STROKE_PT  = 0.8

# Plazas cuyos expedientes se estampan
PLAZAS_STAMP: Set[str] = {"arrecife", "arrecife de lanzarote", "gandia", "telde"}

# Textos que se excluyen del estampado
TEXTOS_EXCLUIR: Set[str] = {"DEMANDA", "PLANTILLAS RESUMEN"}

# ──────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────

def _al(s: str) -> str:
    """ASCII-lower: elimina tildes, minúsculas, colapsa espacios."""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.lower()).strip()

def _load_stamp_log(log_path: str) -> Set[str]:
    """
    Lee stamp_log.csv si existe y devuelve el conjunto de rutas
    que ya fueron estampadas con status=OK o SKIP.
    """
    stamped: Set[str] = set()
    if not os.path.isfile(log_path):
        return stamped
    try:
        with open(log_path, encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f, delimiter=";"):
                if (row.get("status") or "").strip().upper() in ("OK", "SKIP"):
                    ruta = (row.get("ruta") or "").strip()
                    if ruta:
                        stamped.add(ruta)
    except Exception:
        pass
    return stamped


def _exp_from_ruta(ruta: str, in_root: str) -> str:
    """Extrae el nombre del expediente a partir de la ruta completa."""
    try:
        rel   = os.path.relpath(ruta, in_root)
        parts = rel.split(os.sep)
        if len(parts) >= 2 and not parts[0].startswith(".."):
            return parts[0]
    except ValueError:
        pass
    return ""

# ──────────────────────────────────────────
# STAMP
# ──────────────────────────────────────────

def _truncate(text: str, font_name: str, font_size: int, max_w: float) -> str:
    from reportlab.pdfbase.pdfmetrics import stringWidth
    if stringWidth(text, font_name, font_size) <= max_w:
        return text
    t = text
    while t and stringWidth(t + "…", font_name, font_size) > max_w:
        t = t[:-1]
    return (t + "…") if t else "…"


def _make_overlay(page_w: float, page_h: float, text: str):
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from pypdf import PdfReader

    packet  = BytesIO()
    c       = canvas.Canvas(packet, pagesize=(page_w, page_h))
    c.setFont(FONT_NAME, FONT_SIZE)
    c.setLineWidth(BOX_STROKE_PT)

    x      = MARGIN_LEFT_MM * mm
    y_text = page_h - (MARGIN_TOP_MM * mm) - FONT_SIZE

    max_box_w   = page_w - x - (MAX_RIGHT_MM * mm)
    inner_max_w = max(10.0, max_box_w - 2 * BOX_PADDING_PT)

    safe_text = _truncate(text, FONT_NAME, FONT_SIZE, inner_max_w)
    text_w    = stringWidth(safe_text, FONT_NAME, FONT_SIZE)

    box_w = min(max_box_w, text_w + 2 * BOX_PADDING_PT)
    box_h = FONT_SIZE + 2 * BOX_PADDING_PT
    y_box = y_text - BOX_PADDING_PT

    c.rect(x, y_box, box_w, box_h, stroke=1, fill=0)
    c.drawString(x + BOX_PADDING_PT, y_text, safe_text)
    c.save()
    packet.seek(0)
    return PdfReader(packet)


def stamp_pdf(src: str, dst: str, text: str) -> None:
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(src)
    writer = PdfWriter()

    first = reader.pages[0]
    w = float(first.mediabox.width)
    h = float(first.mediabox.height)

    overlay = _make_overlay(w, h, text)
    first.merge_page(overlay.pages[0])

    writer.add_page(first)
    for page in reader.pages[1:]:
        writer.add_page(page)

    dst_dir = os.path.dirname(dst)
    if dst_dir:
        os.makedirs(dst_dir, exist_ok=True)
    with open(dst, "wb") as f:
        writer.write(f)

# ──────────────────────────────────────────
# LECTURA DE INPUTS
# ──────────────────────────────────────────

def load_datatape_plazas(path: str) -> Tuple[Set[str], Dict[str, str]]:
    """
    Lee datatape.xlsx y devuelve:
      - valid_exps : set de Original Contract Numbers en plazas de stamp
      - alias_map  : dict {WCN_con_guiones → OCN} para carpetas nombradas
                     con el Whole Case Number en lugar del OCN
    """
    import openpyxl
    wb  = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws  = wb.active
    exps:      Set[str]        = set()
    alias_map: Dict[str, str]  = {}

    # ── Cabecera → mapa nombre→índice ────────────────────────────
    headers = {}
    header_row = next(ws.iter_rows(values_only=True))
    for i, cell in enumerate(header_row):
        if cell is not None:
            headers[str(cell).strip()] = i

    _COL_EXP  = "Original Contract Number"
    _COL_REF2 = "Whole Case Number"
    _COL_CITY = "City Court"

    def _get(row, col_name):
        idx = headers.get(col_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def _cell_str(val) -> str:
        """Convierte celda a string. Floats enteros → int (evita '40041347800101.0')."""
        if val is None:
            return ""
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val).strip()

    for row in ws.iter_rows(min_row=2, values_only=True):
        exp  = _cell_str(_get(row, _COL_EXP))
        ref2 = _cell_str(_get(row, _COL_REF2))
        city = str(_get(row, _COL_CITY) or "").strip()
        if exp and exp != "None" and _al(city) in PLAZAS_STAMP:
            exps.add(exp)
            if "/" in ref2:
                alias_map[ref2.replace("/", "_")] = exp

    wb.close()
    print(f"[datatape] {len(exps)} expedientes en plazas {sorted(PLAZAS_STAMP)}")
    return exps, alias_map


def load_naming(path: str, valid_exps: Set[str], in_root: str,
                alias_map: Dict[str, str] = None,
                common_dir: str = None,
                ) -> Tuple[Dict[str, List[Dict]], List[Tuple[str, str, str, str]]]:
    """
    Lee documentos.xlsx y devuelve:
      - mapping: exp_folder → [filas] con status=OK, excluyendo DEMANDA y PLANTILLAS RESUMEN
      - common_rows: lista de (asunto_codigo, old_ruta, texto, exp_folder) para filas
        de doc_comun cuyo asunto_codigo pertenece a un expediente de plaza estampado.
    Si la carpeta usa el Whole Case Number con guiones, alias_map normaliza
    el nombre al Original Contract Number.
    """
    import openpyxl
    if alias_map is None:
        alias_map = {}

    mapping: Dict[str, List[Dict]] = {}
    # asunto_codigo → exp_folder (para relacionar filas doc_comun con su expediente)
    asunto_to_exp: Dict[str, str] = {}
    # filas de doc_comun pendientes de resolver (se procesan al final cuando asunto_to_exp esté completo)
    pending_common: List[Dict] = []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(values_only=True))]

    all_rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(raw)}
        all_rows.append(row)
    wb.close()

    for row in all_rows:
        if row.get("status", "").strip().upper() != "OK":
            continue
        texto = row.get("Texto", "").strip()
        if texto in TEXTOS_EXCLUIR:
            continue
        ruta = row.get("ruta", "").strip()
        if not ruta:
            continue

        # Detectar si la ruta apunta a doc_comun
        ruta_norm = ruta.replace("\\", "/")
        if common_dir and ("doc_comun" in ruta_norm or
                           _al(os.path.basename(os.path.dirname(ruta_norm))) == "doc_comun"):
            pending_common.append(row)
            continue

        exp = _exp_from_ruta(ruta, in_root)
        if exp not in valid_exps and exp in alias_map:
            exp = alias_map[exp]
        if exp not in valid_exps:
            continue

        ac = row.get("asunto_codigo", "").strip()
        if ac:
            asunto_to_exp[ac] = exp
        mapping.setdefault(exp, []).append(row)

    # Resolver filas de doc_comun ahora que tenemos asunto_to_exp
    common_rows: List[Tuple[str, str, str, str]] = []
    for row in pending_common:
        ac = row.get("asunto_codigo", "").strip()
        exp = asunto_to_exp.get(ac)
        if not exp:
            continue
        common_rows.append((ac, row.get("ruta", "").strip(), row.get("Texto", "").strip(), exp))

    return mapping, common_rows


def update_documentos_rutas(path: str, ruta_updates: Dict[Tuple[str, str], str]) -> None:
    """
    Reescribe documentos.xlsx actualizando la columna 'ruta' para las filas
    identificadas por (asunto_codigo, old_ruta) con la new_ruta correspondiente.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    try:
        ruta_col  = headers.index("ruta")  + 1   # 1-based
        ac_col    = headers.index("asunto_codigo") + 1
    except ValueError:
        print("[stamp] WARN no se encontró columna 'ruta' o 'asunto_codigo' en documentos.xlsx")
        return

    updated = 0
    for row in ws.iter_rows(min_row=2):
        ac       = str(row[ac_col - 1].value or "").strip()
        old_ruta = str(row[ruta_col - 1].value or "").strip()
        new_ruta = ruta_updates.get((ac, old_ruta))
        if new_ruta:
            row[ruta_col - 1].value = new_ruta
            updated += 1

    wb.save(path)
    print(f"[stamp] documentos.xlsx actualizado: {updated} ruta(s) de doc_comun → carpeta expediente")

# ──────────────────────────────────────────
# PROCESO PRINCIPAL
# ──────────────────────────────────────────

def process(root: str, datatape_path: str, naming_path: str) -> None:

    if not os.path.isfile(datatape_path):
        raise FileNotFoundError(f"No se encuentra el datatape: {datatape_path}")
    if not os.path.isfile(naming_path):
        raise FileNotFoundError(f"No se encuentra el CSV de naming: {naming_path}")

    in_root    = os.path.join(root, "IN")
    common_dir = os.path.join(root, "doc_comun")

    valid_exps, alias_map = load_datatape_plazas(datatape_path)
    if not valid_exps:
        print("Ningún expediente coincide con las plazas indicadas. Nada que hacer.")
        return

    naming, common_rows = load_naming(naming_path, valid_exps, in_root, alias_map, common_dir)
    print(f"[naming]   {sum(len(v) for v in naming.values())} documentos a estampar")
    print(f"[naming]   Expedientes a estampar ({len(valid_exps)}):")
    for exp in sorted(valid_exps):
        n_docs = len(naming.get(exp) or [])
        print(f"  {exp}  →  {n_docs} documento{'s' if n_docs != 1 else ''}")
    print()

    # ── Copiar y estampar documentos comunes en la carpeta del expediente ──
    import shutil
    ruta_updates: Dict[Tuple[str, str], str] = {}   # (ac, old_ruta) → new_ruta

    if common_rows:
        print(f"[doc_comun] {len(common_rows)} doc(s) comunes a copiar y estampar")
    for ac, old_ruta, texto, exp in common_rows:
        fn_orig = os.path.basename(old_ruta)
        fn_copy = f"{exp}_{fn_orig}"
        exp_dir = os.path.join(in_root, exp)
        new_ruta = os.path.join(exp_dir, fn_copy)

        if not os.path.isfile(old_ruta):
            print(f"  [doc_comun] WARN fichero no encontrado: {old_ruta}")
            continue

        os.makedirs(exp_dir, exist_ok=True)
        if not os.path.isfile(new_ruta):
            shutil.copy2(old_ruta, new_ruta)

        tmp = new_ruta + ".stamp_tmp"
        try:
            stamp_pdf(new_ruta, tmp, texto)
            os.replace(tmp, new_ruta)
            ruta_updates[(ac, old_ruta)] = new_ruta
            print(f"  [doc_comun] ✓ {fn_copy}")
        except Exception as e:
            if os.path.exists(tmp):
                os.unlink(tmp)
            print(f"  [doc_comun] FAIL {fn_copy}: {type(e).__name__}: {e}")

    if ruta_updates:
        update_documentos_rutas(naming_path, ruta_updates)
    print()

    log_path     = os.path.join(root, "stamp_log.csv")
    already_done = _load_stamp_log(log_path)
    if already_done:
        print(f"[stamp] stamp_log.csv existente: {len(already_done)} fichero(s) ya estampados previamente.")

    log_rows: List[Dict] = []

    for exp in sorted(valid_exps):
        filas = naming.get(exp)
        if not filas:
            log_rows.append({
                "expediente": exp,
                "fichero":    "",
                "ruta":       "",
                "texto":      "",
                "status":     "WARN",
                "motivo":     "Sin documentos OK en naming_output.csv para este expediente",
            })
            continue

        for fila in filas:
            src        = (fila.get("ruta")  or "").strip()
            stamp_text = (fila.get("Texto") or "").strip()
            fn         = os.path.basename(src)
            tmp        = src + ".stamp_tmp"

            log_row = {
                "expediente": exp,
                "fichero":    fn,
                "ruta":       src,
                "texto":      stamp_text,
                "status":     "",
                "motivo":     "",
            }

            # ── Comprobar si ya fue estampado en una ejecución anterior ──
            if src in already_done:
                log_row["status"] = "SKIP"
                log_row["motivo"] = "Ya estampado previamente"
                print(f"  ↷  {fn}  →  SKIP (ya estampado)")
                log_rows.append(log_row)
                continue

            if not os.path.isfile(src):
                log_row["status"] = "FAIL"
                log_row["motivo"] = f"Fichero no encontrado: {src}"
                log_rows.append(log_row)
                continue

            try:
                stamp_pdf(src, tmp, stamp_text)
                os.replace(tmp, src)   # sustitución atómica
                log_row["status"] = "OK"
                log_row["motivo"] = f"Estampado: '{stamp_text}'"
            except Exception as e:
                if os.path.exists(tmp):
                    os.unlink(tmp)
                log_row["status"] = "FAIL"
                log_row["motivo"] = f"{type(e).__name__}: {e}"

            log_rows.append(log_row)

    # ── Log ──────────────────────────────────────────────────────
    fields = ["expediente", "fichero", "ruta", "texto", "status", "motivo"]
    with open(log_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        w.writeheader()
        w.writerows(log_rows)

    ok   = sum(1 for r in log_rows if r["status"] == "OK")
    skip = sum(1 for r in log_rows if r["status"] == "SKIP")
    warn = sum(1 for r in log_rows if r["status"] == "WARN")
    fail = sum(1 for r in log_rows if r["status"] == "FAIL")
    exps_ok = len({r["expediente"] for r in log_rows if r["status"] == "OK"})

    print(f"Expedientes con documentos OK : {exps_ok} / {len(valid_exps)}")
    print(f"Ficheros  OK                  : {ok}")
    print(f"Ficheros  SKIP (ya estampados): {skip}")
    print(f"Ficheros  WARN                : {warn}  (copiados sin stamp)")
    print(f"Ficheros  FAIL                : {fail}")
    print(f"Log                           : {log_path}")


# ──────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────

if __name__ == "__main__":
    root_dir     = sys.argv[1] if len(sys.argv) > 1 else "/fileserver05/SFTP/KRUK/DEMANDAS"
    datatape     = os.path.join(root_dir, "IN", "datatape.xlsx")
    naming_csv   = os.path.join(root_dir, "documentos.xlsx")
    process(root_dir, datatape, naming_csv)
